import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import time
import io
import os
import tempfile


# File path
EXCEL_FILE = "data.xlsx"
# Defaults to satisfy static analyzer; real values set in each page’s UI
selected_rep = "All"
start_date = end_date = None


# ---------------- Data IO ----------------
# Removing @st.cache_data to avoid caching issues
def load_data():
    sales = pd.read_excel(EXCEL_FILE, sheet_name="SalesLog")
    collections = pd.read_excel(EXCEL_FILE, sheet_name="Collections")
    assignments = pd.read_excel(EXCEL_FILE, sheet_name="Assignments")
    # after: assignments = pd.read_excel(EXCEL_FILE, sheet_name="Assignments")

    # Make sure we have Completed (bool) and TaskStatus (str)
    if "Completed" not in assignments.columns:
        assignments["Completed"] = False
    else:
        assignments["Completed"] = assignments["Completed"].astype(bool)

    if "TaskStatus" not in assignments.columns:
        # default to "Not started" for open rows, "Completed" for completed rows
        assignments["TaskStatus"] = assignments["Completed"].map(lambda x: "Completed" if x else "Not started")
    else:
        assignments["TaskStatus"] = assignments["TaskStatus"].astype(str)

    # NEW: guarantee + normalize the CollectionDate column
    if "CollectionDate" not in collections.columns:
        collections["CollectionDate"] = pd.NaT
    else:
        collections["CollectionDate"] = pd.to_datetime(collections["CollectionDate"], errors="coerce")

    # Ensure numeric
    sales["QuotedPrice"] = pd.to_numeric(sales.get("QuotedPrice", 0), errors="coerce").fillna(0)
    sales["DepositPaid"] = pd.to_numeric(sales.get("DepositPaid", 0), errors="coerce").fillna(0)
    collections["QuoteID"] = pd.to_numeric(collections.get("QuoteID", 0), errors="coerce")
    collections["DepositPaid"] = pd.to_numeric(collections.get("DepositPaid", 0), errors="coerce").fillna(0)

    # Ensure dates
    if "SentDate" in sales.columns:
        sales["SentDate"] = pd.to_datetime(sales["SentDate"], errors="coerce")
        
    # Ensure QuoteID column exists
    if "QuoteID" not in sales.columns:
        sales["QuoteID"] = pd.Series(dtype="int")

    # Normalize QuoteID
    sales["QuoteID"] = pd.to_numeric(sales["QuoteID"], errors="coerce").fillna(0).astype(int)
    collections["QuoteID"] = pd.to_numeric(collections["QuoteID"], errors="coerce").fillna(0).astype(int)

    # 🚫 Drop DepositDue if still present in file
    collections = collections.drop(columns=["DepositDue"], errors="ignore")

    # Ensure Status exists
    if "Status" not in collections.columns:
        collections["Status"] = ""

    return sales, collections, assignments


def save_data(sales, collections, assignments):
    SALES_ORDER = ["QuoteID", "Client", "QuotedPrice", "Status", "SalesRep",
                   "Deposit%", "DepositPaid", "SentDate", "JobType"]
    # No DepositDue here ✅
    COLLECTIONS_ORDER = ["QuoteID", "CollectionDate", "Client", "DepositPaid", "BalanceDue", "Status"]

    # Normalize key columns
    for df in (sales, collections, assignments):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)

    def ensure_and_order(df, desired, fill_defaults=None):
        df = df.copy()
        fill_defaults = fill_defaults or {}
        # make sure desired cols exist
        for col in desired:
            if col not in df.columns:
                df[col] = fill_defaults.get(col, pd.NA)
        # drop DepositDue explicitly if present
        if "DepositDue" in df.columns:
            df = df.drop(columns=["DepositDue"])
        # order: desired first, then extras
        extras = [c for c in df.columns if c not in desired]
        return df[desired + extras]

    sales_to_save = ensure_and_order(
        sales, SALES_ORDER,
        fill_defaults={"Deposit%": 0.0, "DepositPaid": 0.0}
    )
    collections_to_save = ensure_and_order(
        collections, COLLECTIONS_ORDER,
        fill_defaults={"DepositPaid": 0.0, "BalanceDue": 0.0, "Status": ""}
    )
    assignments_to_save = assignments.copy()

    dir_name = os.path.dirname(os.path.abspath(EXCEL_FILE))
    base_name = os.path.basename(EXCEL_FILE)
    tmp_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False, dir=dir_name, prefix=base_name, suffix=".xlsx") as tmp:
            tmp_path = tmp.name

        with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
            sales_to_save.to_excel(writer, sheet_name="SalesLog", index=False)
            collections_to_save.to_excel(writer, sheet_name="Collections", index=False)
            assignments_to_save.to_excel(writer, sheet_name="Assignments", index=False)

            # >>> Apply Excel number formats (currency + date + percent display) <<<
            from openpyxl.utils import get_column_letter

            wb = writer.book
            ws_sales = writer.sheets["SalesLog"]
            ws_col = writer.sheets["Collections"]

            currency_fmt = '"$"#,##0.00'
            date_fmt = 'yyyy-mm-dd'
            percent_literal_fmt = '0.00"%"'  # shows 20 as 20.00% (no scaling)

            def format_column(ws, df, col_name, num_fmt):
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # 1-based
                    col_letter = get_column_letter(col_idx)
                    # skip header at row 1
                    for cell in ws[col_letter][1:]:
                        cell.number_format = num_fmt

            # SalesLog: $ for money, date-only for dates, literal % for Deposit%
            format_column(ws_sales, sales_to_save, "QuotedPrice", currency_fmt)
            format_column(ws_sales, sales_to_save, "DepositPaid", currency_fmt)
            format_column(ws_sales, sales_to_save, "SentDate", date_fmt)
            format_column(ws_sales, sales_to_save, "Deposit%", percent_literal_fmt)

            # Collections: $ for money
            format_column(ws_col, collections_to_save, "DepositPaid", currency_fmt)
            format_column(ws_col, collections_to_save, "BalanceDue", currency_fmt)
            format_column(ws_col, collections_to_save, "CollectionDate", date_fmt)

            

        os.replace(tmp_path, os.path.abspath(EXCEL_FILE))  # atomic replace
        return True
    except PermissionError:
        st.error("⚠️ Could not save data. Please close 'data.xlsx' (Excel might have it open) and try again.")
    except Exception as e:
        st.error(f"💥 Save failed: {type(e).__name__}: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass
    return False



# ---------------- Derivations ----------------
def build_filters_ui(df: pd.DataFrame, key_prefix: str, date_col: str = "SentDate"):
    st.sidebar.subheader("Filters")
    selected_rep, start_date, end_date = "All", None, None

    reps = ["All"]
    if "SalesRep" in df.columns:
        reps += sorted(df["SalesRep"].dropna().unique().tolist())
    selected_rep = st.sidebar.selectbox("Sales Rep", reps, key=f"{key_prefix}_rep")

    if date_col in df.columns and pd.api.types.is_datetime64_any_dtype(df[date_col]):
        valid_dates = df[date_col].dropna()
        if not valid_dates.empty:
            min_date = valid_dates.min().date()
            max_date = valid_dates.max().date()

            picked = st.sidebar.date_input(
                "Select Date Range",
                [min_date, max_date],
                key=f"{key_prefix}_date_range"
            )
            quick = st.sidebar.selectbox(
                "Quick Range",
                [
                    "Choose Quick Range", "Last 7 days", "Last 30 days", "Last 90 days",
                    "This month (MTD)", "Last month", "Year to date (YTD)"
                ],
                index=0,
                key=f"{key_prefix}_quick"
            )

            today = pd.Timestamp.today().normalize()
            first_this_month = today.replace(day=1)

            if quick == "Last 7 days":
                start_date, end_date = (today - pd.Timedelta(days=7)).date(), today.date()
            elif quick == "Last 30 days":
                start_date, end_date = (today - pd.Timedelta(days=30)).date(), today.date()
            elif quick == "Last 90 days":
                start_date, end_date = (today - pd.Timedelta(days=90)).date(), today.date()
            elif quick == "This month (MTD)":
                start_date, end_date = first_this_month.date(), today.date()
            elif quick == "Last month":
                last_month_end = first_this_month - pd.Timedelta(days=1)
                start_date, end_date = last_month_end.replace(day=1).date(), last_month_end.date()
            elif quick == "Year to date (YTD)":
                start_date, end_date = pd.Timestamp(today.year, 1, 1).date(), today.date()
            else:
                # ACCEPT tuple **or** list from date_input
                if isinstance(picked, (list, tuple)) and len(picked) == 2:
                    start_date, end_date = picked[0], picked[1]

    filtered = df.copy()
    if selected_rep != "All" and "SalesRep" in filtered.columns:
        filtered = filtered[filtered["SalesRep"] == selected_rep]

    if start_date and end_date and date_col in filtered.columns:
        # Optional: make end date inclusive even if your datetime has times
        start_ts = pd.to_datetime(start_date)
        end_ts = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
        filtered = filtered[(filtered[date_col] >= start_ts) & (filtered[date_col] <= end_ts)]

    st.session_state[f"{key_prefix}_filters"] = {"rep": selected_rep, "start": start_date, "end": end_date}
    return filtered, selected_rep, start_date, end_date



def apply_saved_filters(df: pd.DataFrame, key_prefix: str, date_col: str = "SentDate"):
    """Re-apply the last chosen filters without re-rendering the sidebar UI."""
    state = st.session_state.get(f"{key_prefix}_filters", {})
    selected_rep = state.get("rep", "All")
    start_date   = state.get("start")
    end_date     = state.get("end")

    filtered = df.copy()
    if selected_rep != "All" and "SalesRep" in filtered.columns:
        filtered = filtered[filtered["SalesRep"] == selected_rep]
    if start_date and end_date and date_col in filtered.columns:
        filtered = filtered[
            (filtered[date_col] >= pd.to_datetime(start_date)) &
            (filtered[date_col] <= pd.to_datetime(end_date))
        ]
    return filtered

def sync_deposit_paid(sales: pd.DataFrame, collections: pd.DataFrame) -> pd.DataFrame:
    """
    Make Sales.DepositPaid = SUM(Collections.DepositPaid) per QuoteID.
    This avoids double counting and keeps Sales as the single source of truth for totals.
    """
    s = sales.copy()
    c = collections.copy()

    # Normalize
    for df in (s, c):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)

    s["QuotedPrice"]   = pd.to_numeric(s.get("QuotedPrice", 0), errors="coerce").fillna(0.0)
    c["DepositPaid"]   = pd.to_numeric(c.get("DepositPaid", 0), errors="coerce").fillna(0.0)

    # Sum every collection (initial + follow-ups) per QuoteID
    sums = c.groupby("QuoteID", dropna=False)["DepositPaid"].sum(min_count=1)
    s["DepositPaid"] = s["QuoteID"].map(sums).fillna(0.0).astype(float)

    # Recompute %
    s["Deposit%"] = s.apply(
        lambda r: round((r["DepositPaid"] / r["QuotedPrice"]) * 100, 2) if r["QuotedPrice"] > 0 else 0.0,
        axis=1
    )
    return s


    # (unused legacy code below intentionally preserved but never reached)
    merged = c.merge(
        s[["QuoteID", "DepositPaid"]].rename(columns={"DepositPaid": "InitialDeposit"}),
        on="QuoteID", how="left"
    )
    merged["InitialDeposit"] = pd.to_numeric(merged["InitialDeposit"], errors="coerce").fillna(0.0)
    raw_sum = merged.groupby("QuoteID")["DepositPaid"].sum()
    has_legacy = (
        (merged["DepositPaid"].round(2) == merged["InitialDeposit"].round(2))
        .groupby(merged["QuoteID"])
        .any()
    )
    initial_map = s.set_index("QuoteID")["DepositPaid"]
    adj_sum = raw_sum - initial_map.where(has_legacy, 0.0)
    adj_sum = adj_sum.reindex(s["QuoteID"].unique(), fill_value=0.0)
    total_paid = initial_map.add(adj_sum, fill_value=0.0)
    s = s.set_index("QuoteID")
    s.loc[total_paid.index, "DepositPaid"] = total_paid.values
    s["Deposit%"] = s.apply(
        lambda r: round((r["DepositPaid"] / r["QuotedPrice"]) * 100, 2) if r["QuotedPrice"] > 0 else 0.0,
        axis=1
    )
    return s.reset_index()



def update_balance_due(sales: pd.DataFrame, collections: pd.DataFrame) -> pd.DataFrame:
    """
    BalanceDue per QuoteID = QuotedPrice - Sales.DepositPaid  (Sales.DepositPaid is TOTAL to date).
    """
    s = sales.copy()
    c = collections.copy()

    # Normalize
    for df in (s, c):
        if "QuoteID" in df.columns:
            df["QuoteID"] = pd.to_numeric(df["QuoteID"], errors="coerce").fillna(0).astype(int)
    s["QuotedPrice"] = pd.to_numeric(s.get("QuotedPrice", 0), errors="coerce").fillna(0.0)
    s["DepositPaid"] = pd.to_numeric(s.get("DepositPaid", 0), errors="coerce").fillna(0.0)

    price_map = s.set_index("QuoteID")["QuotedPrice"]
    paid_map  = s.set_index("QuoteID")["DepositPaid"]
    balance   = (price_map - paid_map).clip(lower=0.0)

    if not c.empty:
        c["BalanceDue"] = c["QuoteID"].map(balance).fillna(0.0).astype(float)
    else:
        if "BalanceDue" not in c.columns:
            c["BalanceDue"] = pd.Series(dtype="float")
    return c



# ---------------- Utils ----------------
# UI display formatting helpers (keep numbers numeric in memory)
def _to_datetime_if_present(df: pd.DataFrame, cols):
    """Coerce the given columns to datetime (in-place-friendly), if they exist."""
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df

def _to_numeric_if_present(df: pd.DataFrame, cols):
    """Coerce the given columns to numeric (in-place-friendly), if they exist."""
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def _fmt_currency_series(s):
    return s.apply(lambda x: f"${x:,.2f}" if pd.notnull(x) else "")

def _fmt_percent_series(s):  # NEW: show 20 -> "20.00%"
    return s.apply(lambda x: f"{float(x):.2f}%" if pd.notnull(x) else "")

# Generate unique QuoteID based on area (Toms River or Manahawkin)
def generate_unique_quote_id(area, sales):
    area = (area or "").strip()
    s = sales.copy()
    s["QuoteID"] = pd.to_numeric(s["QuoteID"], errors="coerce").fillna(0).astype(int)

    if area == "Toms River":
        candidates = s.loc[(s["QuoteID"] >= 1406) & (s["QuoteID"] < 2000), "QuoteID"]
        return (candidates.max() + 1) if not candidates.empty else 1406

    if area == "Manahawkin":
        candidates = s.loc[(s["QuoteID"] >= 2079) & (s["QuoteID"] < 3000), "QuoteID"]
        return (candidates.max() + 1) if not candidates.empty else 2079

    # Fallback range (optional): 3000+
    candidates = s.loc[s["QuoteID"] >= 3000, "QuoteID"]
    return (candidates.max() + 1) if not candidates.empty else 3000


def safe_rerun():
    try:
        st.rerun()
    except Exception:
        st.session_state["force_rerun"] = not st.session_state.get("force_rerun", False)
        st.stop()
        
        # ---------- Inline table editing helpers ----------

def _to_datetime_if_present(df, cols):
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df

def _update_base_by_rowid(base_df: pd.DataFrame, edited_view: pd.DataFrame, cols_to_update, rowid_col="RowID"):
    """
    Push edited values (by row index) back into the original DataFrame.
    'RowID' must be the original base_df index captured before editing.
    """
    updated = base_df.copy()
    e = edited_view.set_index(rowid_col)
    shared = [c for c in cols_to_update if c in updated.columns and c in e.columns]
    updated.loc[e.index, shared] = e[shared]
    return updated


# ---------------- App boot ----------------
# Load data initially
sales, collections, assignments = load_data()

# 👇 add these
sales = sync_deposit_paid(sales, collections)        # roll up initial + collections into Sales.DepositPaid + Deposit%
collections = update_balance_due(sales, collections) # set BalanceDue from Sales totals


# Sidebar
st.sidebar.title("Navigation")
page = st.sidebar.radio("Select Page", ["Dashboard", "Sales Log", "Collections", "Assignments", "View Reports"])

# ---------------- Dashboard ----------------
if page == "Dashboard":
    st.title("📊 Dashboard")
    st.markdown("Here you can review a Quick snapshot of KPIs, Revenue/Payment charts, and Task status.")
    

    # ---- Filters ----
    # Replace your dashboard filter block with:
    filtered_sales, dash_rep, dash_start, dash_end = build_filters_ui(sales, "dash")
    # ...then keep using `filtered_sales` exactly like before


    # ---- Won-only view (copy to avoid SettingWithCopy warnings) ----
    if "Status" in filtered_sales.columns:
        won_sales = filtered_sales[filtered_sales["Status"] == "Won"].copy()
    else:
        won_sales = filtered_sales.iloc[0:0].copy()

    # Ensure numeric QuoteID for joins later
    if "QuoteID" in won_sales.columns:
        won_sales["QuoteID"] = pd.to_numeric(won_sales["QuoteID"], errors="coerce").fillna(0).astype(int)
    if "QuoteID" in collections.columns:
        filtered_collections = collections[collections["QuoteID"].isin(
            won_sales["QuoteID"] if "QuoteID" in won_sales.columns else []
        )].copy()
    else:
        filtered_collections = collections.iloc[0:0].copy()

    # ---- Totals / KPIs ----
 # ---- Totals / KPIs ----

    # Safety: (re)build filtered_sales if it's not in scope
    if 'filtered_sales' not in locals():
        filtered_sales = sales.copy()
        # Reapply filters if those vars exist
        if 'selected_rep' in locals() and selected_rep != "All" and "SalesRep" in filtered_sales.columns:
            filtered_sales = filtered_sales[filtered_sales["SalesRep"] == selected_rep]
        if 'start_date' in locals() and 'end_date' in locals() and start_date and end_date and "SentDate" in filtered_sales.columns:
            filtered_sales = filtered_sales[
                (filtered_sales["SentDate"] >= pd.to_datetime(start_date)) &
                (filtered_sales["SentDate"] <= pd.to_datetime(end_date))
            ]

    # Ensure won_sales and filtered_collections exist
    won_sales = (
        filtered_sales[filtered_sales["Status"] == "Won"].copy()
        if "Status" in filtered_sales.columns else filtered_sales.iloc[0:0].copy()
    )
    if "QuoteID" in collections.columns and "QuoteID" in won_sales.columns:
        won_ids = won_sales["QuoteID"].dropna().astype(int)
        filtered_collections = collections[collections["QuoteID"].isin(won_ids)].copy()
    else:
        filtered_collections = collections.iloc[0:0].copy()

    # Core KPIs
    total_revenue_won = float(won_sales["QuotedPrice"].sum()) if "QuotedPrice" in won_sales.columns else 0.0
    # Average Job Size = Closed $ / Jobs Won
    jobs_won_count = len(won_sales)
    avg_job_size = (total_revenue_won / jobs_won_count) if jobs_won_count > 0 else 0.0


    # Split collections into: initial deposit (first ledger row per QuoteID) vs follow-ups
    if not filtered_collections.empty:
        filtered_collections["DepositPaid"] = pd.to_numeric(
            filtered_collections.get("DepositPaid", 0), errors="coerce"
        ).fillna(0.0)
        initial_by_q = (
            filtered_collections.reset_index()
            .sort_values(["QuoteID", "index"])
            .groupby("QuoteID")["DepositPaid"]
            .first()
        )
        total_deposit_won = float(initial_by_q.sum())  # initial deposits only
        total_collections_won = float(filtered_collections["DepositPaid"].sum() - total_deposit_won)  # follow-ups only
    else:
        total_deposit_won = 0.0
        total_collections_won = 0.0

    # Balance due = price - TOTAL paid (Sales.DepositPaid already equals ledger sum via sync_deposit_paid)
    balance_due_won_jobs = float(
        (won_sales["QuotedPrice"] - won_sales["DepositPaid"]).clip(lower=0).sum()
    ) if {"QuotedPrice","DepositPaid"}.issubset(won_sales.columns) else 0.0

    # Top metrics
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Quotes Sent", len(filtered_sales))
    col2.metric("Jobs Won", len(won_sales))
    col3.metric("Jobs Pending", len(filtered_sales[filtered_sales["Status"] == "Sent"]) if "Status" in filtered_sales.columns else 0)
    col4.metric("Jobs Lost", len(filtered_sales[filtered_sales["Status"] == "Lost"]) if "Status" in filtered_sales.columns else 0)

    col5, col6, col7, col8 = st.columns(4)
    win_rate = (len(won_sales) / len(filtered_sales) * 100) if len(filtered_sales) > 0 else 0.0
    col5.metric("Win Rate %", f"{win_rate:.1f}%")
    col6.metric("Avg Job Size", f"${avg_job_size:,.0f}")
    col7.metric("Closed $", f"${total_revenue_won:,.0f}")
    col8.metric("Deposits Paid", f"${total_deposit_won:,.0f}")   # initial deposits only
    col9, col10, col11, col12 = st.columns(4)
    
    col9.metric("Balance Due", f"${balance_due_won_jobs:,.0f}")
    col10.metric("Total Collected (Collections)", f"${total_collections_won:,.0f}")
    

    # ---- Revenue Breakdown (unchanged logic) ----
    st.subheader("Revenue Breakdown")
    if "Status" in filtered_sales.columns and "QuotedPrice" in filtered_sales.columns:
        revenue_by_status = filtered_sales.groupby("Status")["QuotedPrice"].sum().sort_index()
        total_revenue = float(revenue_by_status.sum())
        if total_revenue > 0:
            fig, ax = plt.subplots()
            bars = ax.bar(
                revenue_by_status.index,
                revenue_by_status.values,
                color=["green" if s == "Won" else "orange" if s == "Lost" else "blue" for s in revenue_by_status.index],
            )
            pct = (revenue_by_status / total_revenue * 100).values
            for bar, p in zip(bars, pct):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(), f"{p:.1f}%", ha="center", va="bottom")
            ax.set_ylabel("Revenue ($)")
            ax.set_title("Revenue Breakdown by Status")
            st.pyplot(fig)
        else:
            st.info("No revenue data available for the selected filters.")
    else:
        st.info("Sales data missing required columns for chart.")

    # ---- Payments Overview (corrected buckets) ----
    st.subheader("Payments Overview")
    if (total_deposit_won > 0) or (total_collections_won > 0) or (balance_due_won_jobs > 0):
        payments_data = {
            "Deposits Paid": total_deposit_won,      # initial only
            "Collections": total_collections_won,    # follow-ups only
            "Balance Due": balance_due_won_jobs,
        }
        total_payments = sum(payments_data.values())
        if total_payments > 0:
            fig2, ax2 = plt.subplots()
            pd.Series(payments_data).plot(kind="bar", ax=ax2, color=["blue", "green", "purple"])
            pct_labels = {k: (v / total_payments * 100) for k, v in payments_data.items()}
            ax2.set_title("Payments Tracking")
            ax2.set_xticklabels([f"{k} ({v:.1f}%)" for k, v in pct_labels.items()], rotation=45)
            st.pyplot(fig2)
        else:
            st.info("No payment data available for the selected filters.")
    else:
        st.info("No payment data available for the selected filters.")

    # ---- Assigned vs Pending Tasks (Won-only) ----
    st.subheader("Task Overview")

    # QuoteIDs that have any assignment rows (unique)
    assigned_tasks = (
        assignments["QuoteID"].dropna().astype(int).unique()
        if "QuoteID" in assignments.columns else []
    )

    # Won jobs under the current filters
    won_quote_ids = (
        won_sales["QuoteID"].dropna().astype(int).unique()
        if "QuoteID" in won_sales.columns else []
    )

    # Pending = won jobs that don't appear in assignments
    pending_tasks = [qid for qid in won_quote_ids if qid not in set(assigned_tasks)]

    assigned_count = int(len(assigned_tasks))
    pending_count  = int(len(pending_tasks))

    # If nothing to show, display a friendly message and skip the chart
    if assigned_count == 0 and pending_count == 0:
        st.info("No task data available for the selected filters.")
    else:
        colA, colB = st.columns(2)
        colA.metric("Assigned Tasks", assigned_count)
        colB.metric("Pending Tasks", pending_count)

        fig3, ax3 = plt.subplots()
        ax3.bar(["Assigned", "Pending"], [assigned_count, pending_count], color=["green", "red"])
        ax3.set_title("Assigned vs Pending Tasks")
        ax3.set_ylabel("Number of Tasks")
        st.pyplot(fig3)

# ---------------- End Dashboard ----------------

# ---------------- Sales Log ----------------
if page == "Sales Log":
    st.title("📝 Sales Log")
    st.markdown("Here you can add New Sales & review Existing Sales.")
    # Right after the title/markdown:
    sl_filtered_sales, sl_rep, sl_start, sl_end = build_filters_ui(sales, "sl")

    # Initialize session state inputs if not present
    if "quoted_price_input" not in st.session_state:
        st.session_state.quoted_price_input = 0.0
    if "deposit_paid_input" not in st.session_state:
        st.session_state.deposit_paid_input = 0.0
    if "sale_added" not in st.session_state:
        st.session_state["sale_added"] = None
    # default place to store the input value for Quote ID
    
        

    # Show success message and ask if want to add another sale
    if st.session_state["sale_added"] is not None:
        new_row = st.session_state["sale_added"]
        st.success(f"✅ Sale added successfully! (Quote ID: {new_row['QuoteID']})")

        add_another = st.radio("Would you like to add another sale?", ["No", "Yes"], index=0)

        if add_another == "Yes":
            # Clear session to show form again immediately
            st.session_state["sale_added"] = None
            st.rerun()
        else:
            st.subheader("Latest Added Sale")
            st.dataframe(pd.DataFrame([new_row]))

    # If no sale added or user selected "Yes" to add another
   # If no sale added or user selected "Yes" to add another
    if st.session_state.get("sale_added") is None:
        st.subheader("Add New Sale")

        # --- Area lives OUTSIDE the form so it can trigger a rerun immediately ---
        area = st.selectbox("Area *", ["Toms River", "Manahawkin"], key="area_select")

        # Suggest a Quote ID whenever Area changes or on first load
        if "quote_id_input" not in st.session_state:
            st.session_state["quote_id_input"] = int(generate_unique_quote_id(area, sales))
            st.session_state["last_area"] = area
        elif st.session_state.get("last_area") != area:
            st.session_state["quote_id_input"] = int(generate_unique_quote_id(area, sales))
            st.session_state["last_area"] = area

        # ---- Single, non-nested form for the rest of the fields ----
        with st.form("add_sale_form"):
            # Editable Quote ID (bound to the same key; no `value=` to avoid the warning)
            st.number_input(
                "Quote ID *",
                min_value=1,
                step=1,
                key="quote_id_input",
                help="Auto-suggested from Area. You can edit it."
            )

            quoted_price = st.number_input("Quoted Price *", min_value=0.0, format="%.2f", key="quoted_price_input")
            deposit_paid = st.number_input("Deposit Paid *", min_value=0.0, format="%.2f", key="deposit_paid_input")
            deposit_pct_preview = round((deposit_paid / quoted_price) * 100, 2) if quoted_price > 0 else 0.0
            st.caption(f"Deposit %: {deposit_pct_preview:.2f}%")

            client = st.text_input("Client *")
            status = st.selectbox("Status *", ["Sent", "Won", "Lost"])
            sales_rep = st.text_input("Sales Rep *")
            sent_date = st.date_input("Sent Date *")
            job_type = st.text_input("Job Type *")

            submitted = st.form_submit_button("Add Sale")

        if submitted:
            quoted_price_val = st.session_state.get("quoted_price_input", 0.0)
            deposit_paid_val = st.session_state.get("deposit_paid_input", 0.0)

            if not client or not sales_rep or not job_type:
                st.error("⚠️ All fields are required. Please fill them in.")
            else:
                # Use Area from the selectbox outside the form
                area = st.session_state["area_select"]

                # Enforce unique Quote ID
                new_id = int(st.session_state["quote_id_input"])
                existing_ids = set(
                    pd.to_numeric(sales.get("QuoteID", pd.Series([], dtype=int)), errors="coerce")
                    .fillna(0).astype(int).tolist()
                )
                if new_id in existing_ids:
                    st.error(
                        f"⚠️ Quote ID {new_id} already exists. "
                        f"Try another (suggested: {generate_unique_quote_id(area, sales)})."
                    )
                    st.stop()

                deposit_pct = round((deposit_paid_val / quoted_price_val) * 100, 2) if quoted_price_val > 0 else 0.0

                new_row = {
                    "QuoteID": new_id,
                    "Client": client,
                    "QuotedPrice": quoted_price_val,
                    "Status": status,
                    "SalesRep": sales_rep,
                    "Deposit%": deposit_pct,
                    "DepositPaid": deposit_paid_val,
                    "SentDate": sent_date,
                    "JobType": job_type
                }
                sales = pd.concat([sales, pd.DataFrame([new_row])], ignore_index=True)

                if deposit_paid_val and deposit_paid_val > 0:
                    collections = pd.concat([collections, pd.DataFrame([{
                        "QuoteID": new_id,
                        "CollectionDate": sent_date,
                        "Client": client,
                        "DepositPaid": deposit_paid_val,
                        "BalanceDue": max(quoted_price_val - deposit_paid_val, 0.0),
                        "Status": "Partially Paid" if deposit_paid_val < quoted_price_val else "Paid",
                    }])], ignore_index=True)

                sales = sync_deposit_paid(sales, collections)
                collections = update_balance_due(sales, collections)
                if save_data(sales, collections, assignments):
                    sales, collections, assignments = load_data()
                    st.session_state["sale_added"] = new_row
                    st.rerun()
                else:
                    st.stop()


    # Always show all sales at the bottom
    st.subheader("All Sales Entries")
    st.caption(f"Rows shown (filtered): {len(sl_filtered_sales)}")

    # Build an edit view using the *base* DataFrame rows so RowID == original index
    sales_edit_idx = sl_filtered_sales.index
    sl_edit_view = sales.loc[sales_edit_idx, ["QuoteID", "Client", "SalesRep", "SentDate", "Status", "JobType", "QuotedPrice"]].copy()
    sl_edit_view.insert(0, "RowID", sl_edit_view.index)  # stable key for writing back

    # Nice editing UX (keeps real dtypes; number/date formats only for UI)
    edited_sl = st.data_editor(
    sl_edit_view,
    key="editor_sales",
    num_rows="fixed",
    disabled=["RowID", "QuoteID"],
    column_config={
        "QuotedPrice": st.column_config.NumberColumn("Quoted Price", format="$%.2f"),
        "SentDate": st.column_config.DateColumn("Sent Date"),
    },
    use_container_width=True,
    hide_index=True,   # 👈 add this
)


    if st.button("Save Sales Table Edits", key="save_sales"):
        # Ensure date types are correct before writing back
        edited_sl = _to_datetime_if_present(edited_sl, ["SentDate"])

        # Push edits back to base `sales` for selected rows/cols
        sales = _update_base_by_rowid(
            sales, edited_sl,
            cols_to_update=["Client", "SalesRep", "SentDate", "Status", "JobType", "QuotedPrice"]
        )

        # Recompute derived fields and persist
        sales = sync_deposit_paid(sales, collections)
        collections = update_balance_due(sales, collections)
        if save_data(sales, collections, assignments):
            st.success("Sales updated.")
            st.rerun()
        else:
            st.stop()

   


# ---------------- Collections ----------------
elif page == "Collections":
    st.title("💰 Collections")
    st.markdown("Search a Quote ID to log follow-up payments and view the live balance due.")
    col_filtered_sales, col_rep, col_start, col_end = build_filters_ui(sales, "col")
    # Won list source:
    won_sales = col_filtered_sales[col_filtered_sales["Status"] == "Won"]

    # Success banner after adding a collection
    if "collection_added_quote" in st.session_state and st.session_state["collection_added_quote"] is not None:
        st.success(f"✅ Collection successfully added for Quote ID {st.session_state['collection_added_quote']}")
        add_another = st.radio("Would you like to add another collection?", ["No", "Yes"], index=0)
        if add_another == "No":
            st.subheader("All Collections Data")
            collections_display = collections.drop(columns=["DepositDue"], errors="ignore").copy()
            for money_col in ("DepositPaid", "BalanceDue"):
                if money_col in collections_display.columns:
                    collections_display[money_col] = _fmt_currency_series(pd.to_numeric(collections_display[money_col], errors="coerce"))
            cols = [c for c in ["QuoteID","Client","DepositPaid","BalanceDue","Status"] if c in collections_display.columns]
            st.dataframe(collections_display[cols] if cols else collections_display)
            st.session_state["collection_added_quote"] = None
            st.stop()
        else:
            st.session_state["collection_added_quote"] = None
            st.session_state.pop("collection_submitted", None)

    # WON jobs only (from the filtered Sales)
    won_sales = col_filtered_sales[col_filtered_sales["Status"] == "Won"]


    # Build the real options only (no dummy first item)
    options = [
        f"{int(row['QuoteID'])} - {row['Client']}"
        for _, row in won_sales.iterrows()
    ]

    # Show only a placeholder in the UI until a value is chosen
    selected = st.selectbox(
        "",  # no visible label
        options,
        index=None,  # nothing preselected
        placeholder="Search QuoteID or Client",
        label_visibility="collapsed",
    )

    if selected:
        selected_quote_id = int(selected.split(" - ")[0])

        # Base info
        sales_row = won_sales[won_sales["QuoteID"] == selected_quote_id]
        client_name = str(sales_row["Client"].values[0]) if not sales_row.empty else ""
        quoted_price = float(sales_row["QuotedPrice"].values[0]) if not sales_row.empty else 0.0

        # 👉 TOTAL paid to date comes straight from Sales (already rolled up by sync_deposit_paid)
        paid_to_date = float(
            sales.loc[sales["QuoteID"] == selected_quote_id, "DepositPaid"].values[0]
        ) if (sales["QuoteID"] == selected_quote_id).any() else 0.0

        # Collections history (raw ledger)
        coll_history = collections[collections["QuoteID"] == selected_quote_id].copy()
        coll_history["DepositPaid"] = pd.to_numeric(coll_history.get("DepositPaid", 0), errors="coerce").fillna(0.0)

        # Summary
        remaining_balance_due = max(quoted_price - paid_to_date, 0.0)
        pay_status = "Paid" if remaining_balance_due == 0 else ("Partially Paid" if paid_to_date > 0 else "Pending")

        st.markdown("### Quote Summary")
        c1, c2, c3 = st.columns(3)
        c1.metric("Quoted Price", f"${quoted_price:,.2f}")
        c2.metric("Paid To Date", f"${paid_to_date:,.2f}")
        c3.metric("Balance Due", f"${remaining_balance_due:,.2f}")
        st.caption(f"Collections ledger sum: ${coll_history['DepositPaid'].sum():,.2f} • Status: {pay_status}")
        # Add new collection
        with st.form("update_collection"):
            st.subheader("Add New Collection")
            st.write(f"Remaining Balance Due: **${remaining_balance_due:,.2f}**")
            collection_date = st.date_input("Collection Date", value=pd.Timestamp.today().date())
            deposit_paid_input = st.number_input("New Collection Amount", value=0.0, min_value=0.0, format="%.2f")
            status_options = ["Pending", "Partially Paid", "Paid", "Overdue"]
            if not coll_history.empty and str(coll_history["Status"].iloc[-1]) in status_options:
                default_status_idx = status_options.index(str(coll_history["Status"].iloc[-1]))
            else:
                default_status_idx = 0
            status_input = st.selectbox("Status", options=status_options, index=default_status_idx)

            submitted = st.form_submit_button("Save Collection")

            if submitted and not st.session_state.get("collection_submitted", False):
                st.session_state["collection_submitted"] = True

                if deposit_paid_input <= 0:
                    st.error("⚠️ Please enter a collection amount greater than 0.00.")
                else:
                    new_row = {
                        "QuoteID": selected_quote_id,
                        "CollectionDate": collection_date,
                        "Client": client_name,
                        "DepositPaid": float(deposit_paid_input),
                        "Status": status_input,
                    }
                    collections = pd.concat([collections, pd.DataFrame([new_row])], ignore_index=True)

                    # Recompute + save (keeps Sales.DepositPaid = TOTAL; Collections.BalanceDue from Sales)
                    sales = sync_deposit_paid(sales, collections)
                    collections = update_balance_due(sales, collections)
                    if save_data(sales, collections, assignments):
                        st.session_state["collection_added_quote"] = selected_quote_id
                        safe_rerun()
                    else:
                        st.stop()
        # Sales info (display-only formatting for $, date-only, and %)
        st.markdown("#### Sales Info")
        if not sales_row.empty:
            sales_view_cols = [c for c in ["QuoteID","Client","QuotedPrice","Status","SalesRep","Deposit%","DepositPaid","SentDate","JobType"] if c in sales_row.columns]
            display_sales_row = sales_row.copy()
            for money_col in ("QuotedPrice", "DepositPaid"):
                if money_col in display_sales_row.columns:
                    display_sales_row[money_col] = _fmt_currency_series(pd.to_numeric(display_sales_row[money_col], errors="coerce"))
            if "Deposit%" in display_sales_row.columns:
                display_sales_row["Deposit%"] = _fmt_percent_series(pd.to_numeric(display_sales_row["Deposit%"], errors="coerce"))
            for date_col in ("SentDate"):
                if date_col in display_sales_row.columns and pd.api.types.is_datetime64_any_dtype(display_sales_row[date_col]):
                    display_sales_row[date_col] = display_sales_row[date_col].dt.date
            st.dataframe(display_sales_row[sales_view_cols])
        else:
            st.info("No Sales row found for this Quote ID.")
        
        # Collections history table with running totals aligned to the Sales total
        st.markdown("#### Collections History")
        if not coll_history.empty:
            base_offset = max(0.0, paid_to_date - coll_history["DepositPaid"].sum())  # covers case where initial deposit isn't in Collections
            coll_history = coll_history.copy()
            coll_history["RunningTotal"] = coll_history["DepositPaid"].cumsum()
            coll_history["TotalPaidAfterThis"] = (base_offset + coll_history["RunningTotal"]).clip(upper=quoted_price)
            coll_history["BalanceAfterThis"] = (quoted_price - coll_history["TotalPaidAfterThis"]).clip(lower=0.0)

            cols_display = [c for c in ["QuoteID","Client","DepositPaid","Status","BalanceDue",
                                        "RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]
                            if c in coll_history.columns or c in ["RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]]
            # format money columns for display
            for money_col in ("DepositPaid", "BalanceDue", "RunningTotal", "TotalPaidAfterThis", "BalanceAfterThis"):
                if money_col in coll_history.columns:
                    coll_history[money_col] = _fmt_currency_series(pd.to_numeric(coll_history[money_col], errors="coerce"))
                    # choose columns (include CollectionDate)
            cols_display = [c for c in ["QuoteID", "CollectionDate", "Client","DepositPaid","Status","BalanceDue",
                                        "RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]
                            if c in coll_history.columns or c in ["RunningTotal","TotalPaidAfterThis","BalanceAfterThis"]]

            # format date for display
            if "CollectionDate" in coll_history.columns and pd.api.types.is_datetime64_any_dtype(coll_history["CollectionDate"]):
                coll_history["CollectionDate"] = coll_history["CollectionDate"].dt.date

            st.dataframe(coll_history[cols_display])
        else:
            st.info("No collections for this Quote ID yet.")

        st.markdown("---")

        

    # Always show all collections 
    st.subheader("All Collections Data")

    # Filter to QuoteIDs present in the (filtered) won sales
    won_ids = (
        won_sales["QuoteID"].dropna().astype(int).unique()
        if "QuoteID" in won_sales.columns else []
    )

    # If nothing to show, bail early with a nice message
    if len(won_ids) == 0:
        st.info("No rows match the current filters.")
    else:
        # Row subset from the *base* df so we can write back by original index
        rows_idx = collections.index[collections["QuoteID"].isin(won_ids)]

        # Build a clean edit view (RAW values, not formatted strings)
        col_edit_view = collections.loc[
            rows_idx,
            ["QuoteID", "CollectionDate", "Client", "DepositPaid", "BalanceDue", "Status"]
        ].copy()

        # Add a visible, named key column (so no blank header)
        col_edit_view.insert(0, "Row", col_edit_view.index.astype(int))

        # Editor (disable derived / key columns)
        edited_col = st.data_editor(
            col_edit_view,
            key="editor_collections",
            num_rows="fixed",                                # add new rows via the form above
            disabled=["Row", "QuoteID", "BalanceDue"],      # BalanceDue is derived
            column_config={
                "CollectionDate": st.column_config.DateColumn("Collection Date"),
                "DepositPaid": st.column_config.NumberColumn("Payment", format="$%.2f"),
                "BalanceDue": st.column_config.NumberColumn("Balance Due", format="$%.2f"),
            },
            use_container_width=True,
            hide_index=True,                                # hide the unnamed pandas index
        )

        st.caption(f"Rows shown (filtered to won jobs): {len(edited_col)}")

        if st.button("Save Collections Table Edits", key="save_collections"):
            # Normalize types coming back from the editor
            edited_col = _to_datetime_if_present(edited_col, ["CollectionDate"])
            edited_col = _to_numeric_if_present(edited_col, ["DepositPaid"])
            
            

            # Write the edited fields back to the *base* collections using the Row key
            for _, r in edited_col.iterrows():
                rid = int(r["Row"])
                for c in ["CollectionDate", "Client", "DepositPaid", "Status"]:
                    if c in collections.columns:
                        collections.at[rid, c] = r[c]

            # Recompute derived fields and persist
            sales = sync_deposit_paid(sales, collections)
            collections = update_balance_due(sales, collections)
            if save_data(sales, collections, assignments):
                st.success("Collections updated.")
                st.rerun()
            else:
                st.stop()
            if st.button("Save Collections Table Edits", key="save_collections"):
            # Normalize types coming back from the editor
                edited_col = _to_datetime_if_present(edited_col, ["CollectionDate"])
                edited_col = _to_numeric_if_present(edited_col, ["DepositPaid"])

            # Write the edited fields back to the *base* collections using the Row key
            for _, r in edited_col.iterrows():
                rid = int(r["Row"])
                for c in ["CollectionDate", "Client", "DepositPaid", "Status"]:
                    if c in collections.columns:
                        collections.at[rid, c] = r[c]

            # Recompute derived fields and persist
            sales = sync_deposit_paid(sales, collections)
            collections = update_balance_due(sales, collections)
            if save_data(sales, collections, assignments):
                st.success("Collections updated.")
                st.rerun()
            else:
                st.stop()

# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# INSERT THE NEW SUMMARY BLOCK *HERE* (still inside the Collections page)
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# --- Completed Jobs (from Assignments) → Balance Due summary ---
    st.markdown("### Completed Jobs — Balance Due")

    # 1) Find completed QuoteIDs in Assignments
    if "Completed" in assignments.columns:
        completed_ids = (
            assignments.loc[assignments["Completed"] == True, "QuoteID"]
            .dropna().astype(int).unique()
        )
    else:
        completed_ids = []

    # 2) Keep in-sync with the current Collections filter (won jobs shown on the page)
    if "QuoteID" in won_sales.columns:
        won_ids_now = set(won_sales["QuoteID"].dropna().astype(int).tolist())
        completed_ids = [qid for qid in completed_ids if qid in won_ids_now]

    # 3) Build the summary from Sales (single row per QuoteID = cleanest BalanceDue)
    if len(completed_ids) == 0:
        st.info("No completed jobs under the current filters.")
    else:
        comp_view = sales.loc[
            sales["QuoteID"].isin(completed_ids),
            ["QuoteID", "Client", "QuotedPrice", "DepositPaid"]
        ].copy()

        # BalanceDue = QuotedPrice - DepositPaid
        comp_view["QuotedPrice"] = pd.to_numeric(comp_view["QuotedPrice"], errors="coerce").fillna(0.0)
        comp_view["DepositPaid"] = pd.to_numeric(comp_view["DepositPaid"], errors="coerce").fillna(0.0)
        comp_view["BalanceDue"] = (comp_view["QuotedPrice"] - comp_view["DepositPaid"]).clip(lower=0.0)

        # Show only the requested columns
        out = comp_view[["QuoteID", "Client", "BalanceDue"]].copy()

        # Sort numerically before formatting
        out = out.sort_values("BalanceDue", ascending=False)

        # Format BalanceDue for display
        out["BalanceDue"] = _fmt_currency_series(out["BalanceDue"])

        st.dataframe(out, use_container_width=True)





# ---------------- Assignments ----------------
elif page == "Assignments":
    st.title("📋 Assignments")
    st.markdown("Assign Pending jobs to crew, set dates/payment/notes, and track pending vs assigned.")

    # ---- Filters (Assignments) ----
    asg_filtered_sales, asg_rep, asg_start, asg_end = build_filters_ui(sales, "asg")
    won_sales = asg_filtered_sales[asg_filtered_sales["Status"] == "Won"]

    # ---- Sets for counts ----
    won_ids_all = won_sales["QuoteID"].dropna().astype(int).tolist() if "QuoteID" in won_sales.columns else []
    assigned_all_ids = set(assignments["QuoteID"].dropna().astype(int).tolist()) if "QuoteID" in assignments.columns else set()
    pending_job_ids = [qid for qid in won_ids_all if qid not in assigned_all_ids]

    # ✅ FIX: open = NOT completed; done = completed
    completed_series = assignments["Completed"] if "Completed" in assignments.columns else pd.Series(False, index=assignments.index)
    asg_open_df = assignments[~completed_series].copy() if not assignments.empty else assignments.copy()
    asg_done_df = assignments[ completed_series].copy() if not assignments.empty else assignments.copy()

    # ---- KPIs ----
    c1, c2, c3 = st.columns(3)
    c1.metric("Assigned Jobs", len(asg_open_df))
    c2.metric("Pending Jobs", len(pending_job_ids))
    c3.metric("Completed", len(asg_done_df))

    st.markdown("---")

    # =================== FORM FIRST ===================
    st.subheader("Assign Tasks")

    if "assigned" not in st.session_state:
        st.session_state.assigned = False

    if st.session_state.assigned:
        assign_another = st.radio("Do you want to assign another task?", options=["Yes", "No"])
        st.session_state.assigned = False  # reset either way

    if not st.session_state.assigned:
        with st.form("assign_task"):
            # Options from *filtered* won jobs
            options = [f"{int(row['QuoteID'])} - {row['Client']}" for _, row in won_sales.iterrows()]

            selected_option = st.selectbox(
                "", options, index=None,
                placeholder="Search Quote ID or Client",
                label_visibility="collapsed",
            )

            if selected_option:
                quote_id = int(selected_option.split(" - ")[0])
                client = selected_option.split(" - ", 1)[1]
            else:
                quote_id = None
                client = None

            crew_member = st.text_input("Crew Member")
            start_date = st.date_input("Start Date")
            end_date = st.date_input("End Date")
            payment = st.number_input("Crew Payment", min_value=0.0, format="%.2f")
            status_choices = ["Not started", "In progress", "On hold", "Completed"]
            task_status = st.selectbox("Task Status", status_choices, index=0)
            notes = st.text_area("Notes")

            days_taken = (end_date - start_date).days if end_date and start_date else 0
            st.caption(f"Days Taken (auto): {days_taken} day(s)")

            submitted = st.form_submit_button("Assign Task")

            if submitted:
                if not crew_member:
                    st.error("⚠️ Please enter the Crew Member name.")
                elif quote_id is None:
                    st.error("⚠️ Please select a valid Quote ID or Client.")
                else:
                    completed_flag = (task_status == "Completed")
                    new_assignment = {
                        "QuoteID": quote_id,
                        "Client": client,
                        "CrewMember": crew_member,
                        "StartDate": start_date,
                        "EndDate": end_date,
                        "Payment": payment,
                        "DaysTaken": days_taken,
                        "Notes": notes,
                        "Completed": completed_flag,
                        "TaskStatus": task_status,
                    }
                    assignments = pd.concat([assignments, pd.DataFrame([new_assignment])], ignore_index=True)

                    if save_data(sales, collections, assignments):
                        st.success(f"Task assigned to {crew_member} for Quote ID {quote_id}.")
                        st.session_state.assigned = True
                        safe_rerun()
                    else:
                        st.stop()

    st.markdown("---")

    # =================== TABS (Tables) ===================
    tab_pending, tab_assigned, tab_completed, tab_all = st.tabs(["Pending Jobs", "Assigned Jobs", "Completed", "All Assignments"])

    # ---- Pending Jobs
    with tab_pending:
        st.subheader("Pending Jobs (Not Yet Assigned)")
        pending_jobs_df = sales[sales["QuoteID"].isin(pending_job_ids)] if "QuoteID" in sales.columns else sales.iloc[0:0]
        if pending_jobs_df.empty:
            st.info("No pending jobs to assign.")
        else:
            base_cols = ["QuoteID","Client","SalesRep","SentDate","QuotedPrice","JobType","Status"]
            show_cols = [c for c in base_cols if c in pending_jobs_df.columns]
            st.dataframe(pending_jobs_df[show_cols] if show_cols else pending_jobs_df, use_container_width=True)

    # ---- Assigned (Open)
    with tab_assigned:
        st.subheader("Assigned Jobs")
        if asg_open_df.empty:
            st.info("No open assignments.")
        else:
            edit_cols = [c for c in ["QuoteID","Client","CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]
                         if c in asg_open_df.columns]
            open_edit_view = asg_open_df[edit_cols].copy()

            edited_open = st.data_editor(
                open_edit_view,
                key="editor_assignments_open",
                num_rows="fixed",
                disabled=["QuoteID","Client"],
                column_config={
                    "StartDate": st.column_config.DateColumn("Start Date"),
                    "EndDate": st.column_config.DateColumn("End Date"),
                    "Payment": st.column_config.NumberColumn("Crew Payment", format="$%.2f"),
                    "TaskStatus": st.column_config.SelectboxColumn(
                        "Task Status",
                        options=["Not started","In progress","On hold","Completed"],
                    ),
                    "Completed": st.column_config.CheckboxColumn("Completed"),
                },
                use_container_width=True,
                hide_index=True,
            )

            if st.button("Save Open Assignments", key="save_assignments_open"):
                edited_open = _to_datetime_if_present(edited_open, ["StartDate","EndDate"])
                edited_open = _to_numeric_if_present(edited_open, ["Payment"])

                for rid, row in edited_open.iterrows():
                    for c in ["CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]:
                        if c in assignments.columns and c in row.index:
                            assignments.at[rid, c] = row[c]
                    # Sync TaskStatus with Completed
                    if bool(assignments.at[rid, "Completed"]):
                        assignments.at[rid, "TaskStatus"] = "Completed"
                    elif assignments.at[rid, "TaskStatus"] == "Completed":
                        assignments.at[rid, "TaskStatus"] = "Not started"

                if {"StartDate","EndDate","DaysTaken"}.issubset(assignments.columns):
                    sd = pd.to_datetime(assignments["StartDate"], errors="coerce")
                    ed = pd.to_datetime(assignments["EndDate"], errors="coerce")
                    assignments["DaysTaken"] = (ed - sd).dt.days.fillna(0).astype(int)

                if save_data(sales, collections, assignments):
                    st.success("Open assignments updated.")
                    st.rerun()
                else:
                    st.stop()

    # ---- Completed
    with tab_completed:
        st.subheader("Completed")
        if asg_done_df.empty:
            st.info("No completed tasks yet.")
        else:
            edit_cols_done = [c for c in ["QuoteID","Client","CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]
                              if c in asg_done_df.columns]
            done_view = asg_done_df[edit_cols_done].copy()

            edited_done = st.data_editor(
                done_view,
                key="editor_assignments_done",
                num_rows="fixed",
                disabled=["QuoteID","Client"],
                column_config={
                    "StartDate": st.column_config.DateColumn("Start Date"),
                    "EndDate": st.column_config.DateColumn("End Date"),
                    "Payment": st.column_config.NumberColumn("Crew Payment", format="$%.2f"),
                    "TaskStatus": st.column_config.SelectboxColumn(
                        "Task Status",
                        options=["Not started","In progress","On hold","Completed"],
                    ),
                    "Completed": st.column_config.CheckboxColumn("Completed"),
                },
                use_container_width=True,
                hide_index=True,
            )

            if st.button("Save Completed Table", key="save_assignments_done"):
                edited_done = _to_datetime_if_present(edited_done, ["StartDate","EndDate"])
                edited_done = _to_numeric_if_present(edited_done, ["Payment"])

                for rid, row in edited_done.iterrows():
                    for c in ["CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]:
                        if c in assignments.columns and c in row.index:
                            assignments.at[rid, c] = row[c]
                    # Sync TaskStatus with Completed
                    if bool(assignments.at[rid, "Completed"]):
                        assignments.at[rid, "TaskStatus"] = "Completed"
                    elif assignments.at[rid, "TaskStatus"] == "Completed":
                        assignments.at[rid, "TaskStatus"] = "Not started"

                if {"StartDate","EndDate","DaysTaken"}.issubset(assignments.columns):
                    sd = pd.to_datetime(assignments["StartDate"], errors="coerce")
                    ed = pd.to_datetime(assignments["EndDate"], errors="coerce")
                    assignments["DaysTaken"] = (ed - sd).dt.days.fillna(0).astype(int)

                if save_data(sales, collections, assignments):
                    st.success("Completed table updated.")
                    st.rerun()
                else:
                    st.stop()

    # ---- All (everything, one editable table)
    with tab_all:
        st.subheader("All Assignments (Editable)")
        if assignments.empty:
            st.info("No assignments yet.")
        else:
            all_cols = [c for c in ["QuoteID","Client","CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]
                        if c in assignments.columns]
            all_view = assignments[all_cols].copy()

            edited_all = st.data_editor(
                all_view,
                key="editor_assignments_all",
                num_rows="fixed",
                disabled=["QuoteID","Client"],
                column_config={
                    "StartDate": st.column_config.DateColumn("Start Date"),
                    "EndDate": st.column_config.DateColumn("End Date"),
                    "Payment": st.column_config.NumberColumn("Crew Payment", format="$%.2f"),
                    "TaskStatus": st.column_config.SelectboxColumn(
                        "Task Status",
                        options=["Not started","In progress","On hold","Completed"],
                    ),
                    "Completed": st.column_config.CheckboxColumn("Completed"),
                },
                use_container_width=True,
                hide_index=True,
            )

            if st.button("Save All Assignments", key="save_assignments_all"):
                edited_all = _to_datetime_if_present(edited_all, ["StartDate","EndDate"])
                edited_all = _to_numeric_if_present(edited_all, ["Payment"])

                for rid, row in edited_all.iterrows():
                    for c in ["CrewMember","StartDate","EndDate","Payment","TaskStatus","Notes","Completed"]:
                        if c in assignments.columns and c in row.index:
                            assignments.at[rid, c] = row[c]
                    # Sync TaskStatus with Completed
                    if bool(assignments.at[rid, "Completed"]):
                        assignments.at[rid, "TaskStatus"] = "Completed"
                    elif assignments.at[rid, "TaskStatus"] == "Completed":
                        assignments.at[rid, "TaskStatus"] = "Not started"

                if {"StartDate","EndDate","DaysTaken"}.issubset(assignments.columns):
                    sd = pd.to_datetime(assignments["StartDate"], errors="coerce")
                    ed = pd.to_datetime(assignments["EndDate"], errors="coerce")
                    assignments["DaysTaken"] = (ed - sd).dt.days.fillna(0).astype(int)

                if save_data(sales, collections, assignments):
                    st.success("Assignments updated.")
                    st.rerun()
                else:
                    st.stop()


# ---------------- View Reports ----------------
if page == "View Reports":
    st.title("📊 View Reports")
    st.markdown("Here you can view and download reports for Sales Log, Collections, and Assignments.")
    rpt_filtered_sales, rpt_rep, rpt_start, rpt_end = build_filters_ui(sales, "rpt")

    sales, collections, assignments = load_data()

    report_tabs = st.radio("Select a Report to View", ("Sales Log", "Collections", "Assignments"))

    # Tab bodies:
    if report_tabs == "Sales Log":
        st.subheader("Sales Log (Filtered)")
        st.write(rpt_filtered_sales)
    elif report_tabs == "Collections":
        st.subheader("Collections (Filtered to matching Sales)")
        rpt_ids = rpt_filtered_sales["QuoteID"].dropna().astype(int).unique() if "QuoteID" in rpt_filtered_sales.columns else []
        st.write(collections[collections["QuoteID"].isin(rpt_ids)])
    elif report_tabs == "Assignments":
        st.subheader("Assignments (Filtered to matching Sales)")
        rpt_ids = rpt_filtered_sales["QuoteID"].dropna().astype(int).unique() if "QuoteID" in rpt_filtered_sales.columns else []
        st.write(assignments[assignments["QuoteID"].isin(rpt_ids)])
        # Download button for unified report (Excel file with three tabs)
        st.subheader("Download Unified Report")
        st.markdown("Click the button below to download the report with Sales Log, Collections, and Assignments as separate tabs in an Excel file.")

    def create_excel_report(sales, collections, assignments):
        excel_file = io.BytesIO()
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            sales.to_excel(writer, sheet_name="Sales Log", index=False)
            collections.to_excel(writer, sheet_name="Collections", index=False)
            assignments.to_excel(writer, sheet_name="Assignments", index=False)
        excel_file.seek(0)
        return excel_file

    if st.button("Download Unified Report"):
        excel_file = create_excel_report(sales, collections, assignments)
        st.download_button(
            label="Download Unified Report (Excel)",
            data=excel_file,
            file_name="unified_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
